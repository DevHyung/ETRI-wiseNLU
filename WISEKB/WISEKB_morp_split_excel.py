from openpyxl import  load_workbook
from openpyxl import Workbook
if __name__== "__main__":
    tEx = load_workbook(filename='대화.xlsx')
    sheet1 = tEx.active
    dataList = []
    for i in sheet1.rows:
        origin = i[1].value
        result_morp = i[2].value
        dataList.append([origin, result_morp])
    print(">>> 총 {} 개 데이터 입력완료".format(len(dataList)))

    book = Workbook()
    sheet = book.active
    sheet.title = 'result'
    sheet.column_dimensions['A'].width = 45
    sheet.column_dimensions['B'].width = 40
    sheet.column_dimensions['C'].width = 10
    for data in dataList:
        sheet.append([data[0],data[1],len(data[1].split(' '))] + data[1].split(' '))
        sheet.append([])


    book.save("분석.xlsx")


